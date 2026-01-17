from openpyxl.styles import Font, Alignment, Border, Side
from backend.reports.charts import generate_percentage_band_chart

bold = Font(bold=True)
center = Alignment(horizontal="center")
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_table(ws, r1, c1, r2, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = border


def generate_analysis_sheet(ws, summary, toppers, best_subjects, subject_perf):

    # ========== RESULT AT A GLANCE (HORIZONTAL) ==========
    ws["B2"] = "RESULT AT A GLANCE"
    ws["B2"].font = bold

    headers = [
        "Total Students","Passed","Compartment","Failed",
        "Topper %","Average %","95%+ Students"
    ]
    values = [
        summary.total_students,
        summary.passed,
        summary.compartment,
        summary.failed,
        summary.topper_percentage,
        summary.average_percentage,
        len(toppers)
    ]

    col = 2
    for h in headers:
        ws.cell(row=4, column=col, value=h).font = bold
        ws.cell(row=4, column=col).alignment = center
        col += 1

    col = 2
    for v in values:
        ws.cell(row=5, column=col, value=v).alignment = center
        col += 1

    style_table(ws,4,2,5,8)

    # ========== TOPPERS ==========
    ws["B8"] = "TOPPERS"
    ws["B8"].font = bold

    heads = ["Rank","Name","Percentage"]
    for i,h in enumerate(heads,2):
        ws.cell(row=10,column=i,value=h).font = bold

    r = 11
    for i,t in enumerate(toppers,1):
        ws.cell(row=r,column=2,value=i)
        ws.cell(row=r,column=3,value=t["name"])
        ws.cell(row=r,column=4,value=t["percentage"])
        r+=1

    style_table(ws,10,2,r-1,4)

    # ========== SUBJECT PERFORMANCE ==========
    start = r + 2
    ws[f"B{start}"] = "SUBJECT WISE PERFORMANCE"
    ws[f"B{start}"].font = bold

    heads = ["Subject","Appeared","Pass","Fail","Pass %"]
    for i,h in enumerate(heads,2):
        ws.cell(row=start+2,column=i,value=h).font = bold

    rr = start+3
    for sub,data in subject_perf.items():
        ws.cell(row=rr,column=2,value=sub)
        ws.cell(row=rr,column=3,value=data["appeared"])
        ws.cell(row=rr,column=4,value=data["pass"])
        ws.cell(row=rr,column=5,value=data["fail"])
        ws.cell(row=rr,column=6,value=data["pass_percentage"])
        rr+=1

    style_table(ws,start+2,2,rr-1,6)

    # ========== BEST IN SUBJECT ==========
    ws[f"H{start}"] = "BEST IN SUBJECT"
    ws[f"H{start}"].font = bold

    heads = ["Subject","Student(s)","Marks"]
    for i,h in enumerate(heads,8):
        ws.cell(row=start+2,column=i,value=h).font = bold

    rb = start+3
    for sub,data in best_subjects.items():
        ws.cell(row=rb,column=8,value=sub)
        ws.cell(row=rb,column=9,value=", ".join(data["students"]))
        ws.cell(row=rb,column=10,value=data["percentage"])
        rb+=1

    style_table(ws,start+2,8,rb-1,10)

    # ========== PIE CHART ==========
    generate_percentage_band_chart(ws, summary.performance_bands, rb+2)
