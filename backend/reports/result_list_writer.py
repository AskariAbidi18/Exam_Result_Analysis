from backend.core.models import Student
from typing import List
from openpyxl.styles import Font, Border, Side, Alignment

bold = Font(bold=True)
thin = Side(style="thin")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

center = Alignment(horizontal="center")


def generate_result_sheet(
    ws,
    students: List[Student],
    exam_type: str,
    start_row: int = 2
):

    # -----------------------------
    # HEADERS
    # -----------------------------

    headers = [
        "Roll No",
        "Name",
        "Gender"
    ]

    for i in range(1, 7):

        headers.append(f"Sub{i} Code")
        headers.append(f"Sub{i} Marks")

        # grades only for class 12
        if exam_type == "12":
            headers.append(f"Sub{i} Grade")

    headers += [
        "Total Marks",
        "Percentage",
        "Result"
    ]

    # -----------------------------
    # WRITE HEADERS
    # -----------------------------

    for col, h in enumerate(headers, start=1):

        cell = ws.cell(
            row=1,
            column=col,
            value=h
        )

        cell.font = bold
        cell.border = border
        cell.alignment = center

    ws.freeze_panes = "A2"

    # -----------------------------
    # DATA
    # -----------------------------

    row = start_row

    for student in students:

        col = 1

        ws.cell(
            row=row,
            column=col,
            value=student.roll_no
        )
        col += 1

        ws.cell(
            row=row,
            column=col,
            value=student.name
        )
        col += 1

        ws.cell(
            row=row,
            column=col,
            value=student.gender
        )
        col += 1

        # -----------------------------
        # SUBJECTS
        # -----------------------------

        for subject in student.subjects:

            ws.cell(
                row=row,
                column=col,
                value=subject.subject_code
            )
            col += 1

            ws.cell(
                row=row,
                column=col,
                value=subject.marks
            )
            col += 1

            # grades only for class 12
            if exam_type == "12":

                ws.cell(
                    row=row,
                    column=col,
                    value=subject.grade
                )
                col += 1

        # -----------------------------
        # EMPTY PADDING
        # -----------------------------

        remaining = 6 - len(student.subjects)

        for _ in range(remaining):

            ws.cell(row=row, column=col, value="")
            col += 1

            ws.cell(row=row, column=col, value="")
            col += 1

            if exam_type == "12":
                ws.cell(row=row, column=col, value="")
                col += 1

        # -----------------------------
        # TOTAL / PERCENTAGE / RESULT
        # -----------------------------

        ws.cell(
            row=row,
            column=col,
            value=student.total_marks()
        )
        col += 1

        ws.cell(
            row=row,
            column=col,
            value=student.percentage()
        )
        col += 1

        ws.cell(
            row=row,
            column=col,
            value=student.result_status
        )

        # -----------------------------
        # BORDERS
        # -----------------------------

        for c in range(1, col + 1):

            ws.cell(
                row=row,
                column=c
            ).border = border

        row += 1
        