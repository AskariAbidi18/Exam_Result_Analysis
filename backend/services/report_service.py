from openpyxl import Workbook
from openpyxl import load_workbook

from backend.core.parser import parse_raw_data
from backend.core.calculator import generate_result_summary
from backend.core.analysis import (
    get_toppers,
    get_best_in_subject,
    get_subject_wise_performance
)
from backend.reports.excel_writer import generate_analysis_sheet
from backend.reports.result_list_writer import generate_result_sheet


def generate_full_report(raw_file_path, template_path, output_path):
    students = parse_raw_data(raw_file_path)

    summary = generate_result_summary(students)
    toppers = get_toppers(students)
    best_subjects = get_best_in_subject(students)
    subject_performance = get_subject_wise_performance(students)

    # Step 1: create Result sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    generate_result_sheet(ws, students)
    wb.save(output_path)

    # Step 2: append Analysis sheet from template
    analysis_wb = load_workbook(template_path)
    analysis_ws = analysis_wb.active
    analysis_ws.title = "Analysis"

    wb = load_workbook(output_path)
    wb._add_sheet(analysis_ws)
    wb.save(output_path)

    # Step 3: fill Analysis sheet
    generate_analysis_sheet(
        template_path=output_path,
        output_path=output_path,
        summary=summary,
        toppers=toppers,
        best_subjects=best_subjects,
        subject_performance=subject_performance
    )
