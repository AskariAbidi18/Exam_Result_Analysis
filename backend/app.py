from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import uuid, os, shutil
from pathlib import Path

from backend.core.parser import parse_raw_data
from backend.core.calculator import generate_result_summary
from backend.core.analysis import (
    get_toppers,
    get_best_in_subject,
    get_subject_wise_performance
)
from backend.reports.excel_writer import generate_analysis_sheet
from backend.reports.result_list_writer import generate_result_sheet
from backend.main import auto_width

from openpyxl import Workbook

app = FastAPI()

# CORS (important for frontend)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "data/raw"
OUTPUT_DIR = "data/output"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


@app.post("/generate-report")
async def generate_report(file: UploadFile = File(...)):
    try:
        # Validate file type
        if not file.filename.endswith('.txt'):
            raise HTTPException(status_code=400, detail="Only .txt files are supported")

        # Read file content directly into memory
        content = await file.read()
        
        # Save to temporary file for parsing (required by parse_raw_data)
        import tempfile
        with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.txt') as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        print(f"Processing file: {file.filename}")

        try:
            # Run pipeline
            students = parse_raw_data(tmp_path)
            print(f"Parsed {len(students)} students")

            wb = Workbook()
            wb.remove(wb.active)

            result_ws = wb.create_sheet("Result")
            generate_result_sheet(result_ws, students)

            analysis_ws = wb.create_sheet("Analysis")

            summary = generate_result_summary(students)
            toppers = get_toppers(students)
            best_subjects = get_best_in_subject(students)
            subject_perf = get_subject_wise_performance(students)

            generate_analysis_sheet(
                analysis_ws,
                summary,
                toppers,
                best_subjects,
                subject_perf
            )

            auto_width(result_ws)
            auto_width(analysis_ws)

            # Save to BytesIO instead of file
            from io import BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            print(f"Excel file created in memory, size: {len(excel_buffer.getvalue())} bytes")

            # Return file as streaming response
            return StreamingResponse(
                excel_buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=Final_Report.xlsx",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
        
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")


@app.get("/")
async def root():
    return {"message": "Board Exam Result Analyzer API is running"}


@app.get("/health")
async def health_check():
    return {"status": "ok"}