import uuid
from openpyxl import Workbook

from backend.core.parser import parse_raw_data
from backend.core.calculator import generate_result_summary
from backend.core.analysis import (
    get_toppers,
    get_best_in_subject,
    get_subject_wise_performance
)
from backend.reports.excel_writer import generate_analysis_sheet
from backend.reports.result_list_writer import generate_result_sheet


RAW_FILE = "data/raw/raw_data.txt"
OUTPUT_FILE = f"data/output/ResultAnalysis.xlsx"

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3



def main():

    # 1. Parse raw data
    students = parse_raw_data(RAW_FILE)

    # 2. Create new workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 3. Create RESULT sheet
    result_ws = wb.create_sheet(title="Result")
    generate_result_sheet(result_ws, students)

    # 4. Create ANALYSIS sheet
    analysis_ws = wb.create_sheet(title="Analysis")

    summary = generate_result_summary(students)
    toppers = get_toppers(students)
    best_subjects = get_best_in_subject(students)
    subject_perf = get_subject_wise_performance(students)

    generate_analysis_sheet(
        analysis_ws,
        summary,
        toppers,
        best_subjects,
        subject_perf
    )

    auto_width(result_ws)
    auto_width(analysis_ws)


    # 5. Save output
    wb.save(OUTPUT_FILE)
    print("✅ FINAL REPORT GENERATED:", OUTPUT_FILE)


if __name__ == "__main__":
    main()
